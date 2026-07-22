from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
xlsx = ROOT / "St. Andrew's Point Pioneer Cemetery Data.xlsx"

wb = load_workbook(xlsx, data_only=True)
print("Sheets:", wb.sheetnames)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\nSheet: {sheet} | rows={ws.max_row} cols={ws.max_column}")
    for r in range(1, min(16, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 14) + 1)]
        print(r, vals)
