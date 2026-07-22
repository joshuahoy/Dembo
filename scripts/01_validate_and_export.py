import csv
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "St. Andrew's Point Pioneer Cemetery Data.xlsx"
DATA_DIR = ROOT / "data"
DOCS_DATA_DIR = ROOT / "docs" / "data"
VALIDATION_DIR = ROOT / "validation"

PEI_BOUNDS = {
    "lat_min": 45.5,
    "lat_max": 47.2,
    "lon_min": -64.6,
    "lon_max": -61.5,
}


def to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\u00a0", "").strip()
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def solve_3x3(a: List[List[float]], b: List[float]) -> List[float]:
    m = [row[:] + [rhs] for row, rhs in zip(a, b)]
    n = 3

    for col in range(n):
        pivot = max(range(col, n), key=lambda r: abs(m[r][col]))
        if abs(m[pivot][col]) < 1e-12:
            raise ValueError("Singular matrix while solving georeference transform")
        if pivot != col:
            m[col], m[pivot] = m[pivot], m[col]

        pivot_val = m[col][col]
        for c in range(col, n + 1):
            m[col][c] /= pivot_val

        for r in range(n):
            if r == col:
                continue
            factor = m[r][col]
            for c in range(col, n + 1):
                m[r][c] -= factor * m[col][c]

    return [m[i][n] for i in range(n)]


def fit_plane(points: List[Tuple[float, float, float]]) -> Tuple[float, float, float]:
    # target = c0 + c1 * northing + c2 * easting
    s1 = len(points)
    sy = sum(p[0] for p in points)
    sx = sum(p[1] for p in points)
    syy = sum(p[0] * p[0] for p in points)
    sxx = sum(p[1] * p[1] for p in points)
    syx = sum(p[0] * p[1] for p in points)

    st = sum(p[2] for p in points)
    syt = sum(p[0] * p[2] for p in points)
    sxt = sum(p[1] * p[2] for p in points)

    a = [
        [s1, sy, sx],
        [sy, syy, syx],
        [sx, syx, sxx],
    ]
    b = [st, syt, sxt]
    c0, c1, c2 = solve_3x3(a, b)
    return c0, c1, c2


def is_in_pei(lat: float, lon: float) -> bool:
    return (
        PEI_BOUNDS["lat_min"] <= lat <= PEI_BOUNDS["lat_max"]
        and PEI_BOUNDS["lon_min"] <= lon <= PEI_BOUNDS["lon_max"]
    )


def build_feature(record: Dict) -> Dict:
    props = {
        "source_sheet": record["source_sheet"],
        "point_type": record["point_type"],
        "point_id": record["point_id"],
        "northing_y": record["northing_y"],
        "easting_x": record["easting_x"],
        "category": record.get("category"),
        "grid_context": record.get("grid_context"),
        "is_anchor": record.get("is_anchor", False),
    }
    if record.get("elevation") is not None:
        props["elevation"] = record["elevation"]
    return {
        "type": "Feature",
        "geometry": {
            "type": "Point",
            "coordinates": [record["longitude"], record["latitude"]],
        },
        "properties": props,
    }


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_DATA_DIR.mkdir(parents=True, exist_ok=True)
    VALIDATION_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(XLSX_PATH, data_only=True)

    coord_ws = wb["Coordinate Points"]
    anchors = []
    anchor_records = []
    missing_coord_latlon = 0

    for row in range(2, coord_ws.max_row + 1):
        point_id = coord_ws.cell(row, 1).value
        northing = to_float(coord_ws.cell(row, 2).value)
        easting = to_float(coord_ws.cell(row, 3).value)
        lat = to_float(coord_ws.cell(row, 4).value)
        lon = to_float(coord_ws.cell(row, 5).value)
        elev = to_float(coord_ws.cell(row, 6).value)
        context = coord_ws.cell(row, 7).value

        if northing is None and easting is None and lat is None and lon is None:
            continue

        if lat is not None and lon is not None and northing is not None and easting is not None:
            anchors.append((northing, easting, lat, lon))

        if lat is None or lon is None:
            missing_coord_latlon += 1
            continue

        anchor_records.append(
            {
                "source_sheet": "Coordinate Points",
                "point_type": "anchor",
                "point_id": point_id,
                "northing_y": northing,
                "easting_x": easting,
                "latitude": lat,
                "longitude": lon,
                "elevation": elev,
                "grid_context": context,
                "category": None,
                "is_anchor": True,
            }
        )

    if len(anchors) < 3:
        raise RuntimeError("Need at least 3 anchor points with X/Y and lat/lon to georeference data")

    lat_fit = fit_plane([(a[0], a[1], a[2]) for a in anchors])
    lon_fit = fit_plane([(a[0], a[1], a[3]) for a in anchors])

    def project(northing: float, easting: float) -> Tuple[float, float]:
        lat = lat_fit[0] + lat_fit[1] * northing + lat_fit[2] * easting
        lon = lon_fit[0] + lon_fit[1] * northing + lon_fit[2] * easting
        return lat, lon

    records = []
    records.extend(anchor_records)

    def parse_sheet(sheet_name: str, point_type: str, category_col: Optional[int] = None) -> None:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            point_id = ws.cell(row, 1).value
            northing = to_float(ws.cell(row, 2).value)
            easting = to_float(ws.cell(row, 3).value)

            if point_id is None and northing is None and easting is None:
                continue
            if northing is None or easting is None:
                continue

            lat, lon = project(northing, easting)
            category = ws.cell(row, category_col).value if category_col else None

            records.append(
                {
                    "source_sheet": sheet_name,
                    "point_type": point_type,
                    "point_id": point_id,
                    "northing_y": northing,
                    "easting_x": easting,
                    "latitude": lat,
                    "longitude": lon,
                    "elevation": None,
                    "grid_context": None,
                    "category": category,
                    "is_anchor": False,
                }
            )

    parse_sheet("Surface Stones ", "surface")
    parse_sheet("Subterranean Points", "subterranean", category_col=7)

    out_of_bounds = [r for r in records if not is_in_pei(r["latitude"], r["longitude"])]

    # Deduplicate by source/type/id/y/x to guard accidental duplicates.
    unique = []
    seen = set()
    for r in records:
        key = (
            r["source_sheet"],
            r["point_type"],
            r["point_id"],
            round(r["northing_y"] if r["northing_y"] is not None else -9999, 6),
            round(r["easting_x"] if r["easting_x"] is not None else -9999, 6),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(r)

    csv_path = DATA_DIR / "cemetery_clean.csv"
    csv_fields = [
        "source_sheet",
        "point_type",
        "point_id",
        "northing_y",
        "easting_x",
        "latitude",
        "longitude",
        "elevation",
        "category",
        "grid_context",
        "is_anchor",
    ]

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fields)
        writer.writeheader()
        for r in unique:
            writer.writerow(r)

    geojson = {
        "type": "FeatureCollection",
        "features": [build_feature(r) for r in unique],
    }

    geojson_path = DATA_DIR / "cemetery_clean.geojson"
    with geojson_path.open("w", encoding="utf-8") as f:
        json.dump(geojson, f, ensure_ascii=False, indent=2)

    docs_geojson_path = DOCS_DATA_DIR / "cemetery_clean.geojson"
    with docs_geojson_path.open("w", encoding="utf-8") as f:
        json.dump(geojson, f, ensure_ascii=False, indent=2)

    lat_residuals = []
    lon_residuals = []
    for n, e, lat_true, lon_true in anchors:
        lat_pred, lon_pred = project(n, e)
        lat_residuals.append(abs(lat_pred - lat_true))
        lon_residuals.append(abs(lon_pred - lon_true))

    report = []
    report.append("St. Andrew's Point Coordinate Validation Report")
    report.append("=" * 48)
    report.append(f"Source workbook: {XLSX_PATH.name}")
    report.append("")
    report.append("Anchor fit summary")
    report.append(f"- Anchor points used: {len(anchors)}")
    report.append(f"- Latitude fit coeffs: c0={lat_fit[0]:.10f}, c1={lat_fit[1]:.10f}, c2={lat_fit[2]:.10f}")
    report.append(f"- Longitude fit coeffs: c0={lon_fit[0]:.10f}, c1={lon_fit[1]:.10f}, c2={lon_fit[2]:.10f}")
    report.append(f"- Max |lat residual| on anchors: {max(lat_residuals):.10f} deg")
    report.append(f"- Max |lon residual| on anchors: {max(lon_residuals):.10f} deg")
    report.append("")
    report.append("Dataset summary")
    report.append(f"- Total exported points: {len(unique)}")
    report.append(f"- Anchor points exported: {sum(1 for r in unique if r['is_anchor'])}")
    report.append(f"- Surface points exported: {sum(1 for r in unique if r['point_type'] == 'surface')}")
    report.append(f"- Subterranean points exported: {sum(1 for r in unique if r['point_type'] == 'subterranean')}")
    report.append(f"- Coordinate sheet rows missing lat/lon: {missing_coord_latlon}")
    report.append(f"- Out-of-PEI bound points: {len(out_of_bounds)}")

    if out_of_bounds:
        report.append("")
        report.append("Out-of-bounds sample")
        for r in out_of_bounds[:10]:
            report.append(
                f"- {r['source_sheet']} id={r['point_id']} lat={r['latitude']:.6f} lon={r['longitude']:.6f}"
            )

    report_path = VALIDATION_DIR / "coordinate_report.txt"
    report_path.write_text("\n".join(report), encoding="utf-8")

    print(f"Wrote: {csv_path}")
    print(f"Wrote: {geojson_path}")
    print(f"Wrote: {docs_geojson_path}")
    print(f"Wrote: {report_path}")
    print(f"Exported points: {len(unique)}")


if __name__ == "__main__":
    main()
